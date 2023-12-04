import pandas as pd

from openpyxl import load_workbook


sheet_list = load_workbook("template_upload.xlsx", read_only=True).sheetnames
# print(sheet_list)

def calculate_aqi_2_5(median_value):
    if 0.0 <= median_value <= 12.0:
        aqi = ((50 - 0) / (12.0 - 0.0)) * (median_value - 0.0) + 0
    elif 12.1 <= median_value <= 35.4:
        aqi = ((100 - 51) / (35.4 - 12.1)) * (median_value - 12.1) + 51
    elif 35.5 <= median_value <= 55.4:
        aqi = ((150 - 101) / (55.4 - 35.5)) * (median_value - 35.5) + 101
    elif 55.5 <= median_value <= 150.4:
        aqi = ((200 - 151) / (150.4 - 55.5)) * (median_value - 55.5) + 151
    elif 150.5 <= median_value <= 250.4:
        aqi = ((300 - 201) / (250.4 - 150.5)) * (median_value - 150.5) + 201
    elif 250.5 <= median_value <= 350.4:
        aqi = ((400 - 301) / (350.4 - 250.5)) * (median_value - 250.5) + 301
    elif 350.5 <= median_value <= 500.4:
        aqi = ((500 - 401) / (500.4 - 350.5)) * (median_value - 350.5) + 401
    else:
        aqi = 500  # Beyond the AQI scale

    return round(aqi)

def calculate_aqi_10(median_value):
    if 0.0 <= median_value <= 54.0:
        aqi = ((50 - 0) / (54.0 - 0.0)) * (median_value - 0.0) + 0
    elif 55.0 <= median_value <= 154.0:
        aqi = ((100 - 51) / (154.0 - 55.0)) * (median_value - 55.0) + 51
    elif 155.0 <= median_value <= 254.0:
        aqi = ((150 - 101) / (254.0 - 155.0)) * (median_value - 155.0) + 101
    elif 255.0 <= median_value <= 354.0:
        aqi = ((200 - 151) / (354.0 - 255.0)) * (median_value - 255.0) + 151
    elif 355.0 <= median_value <= 424.0:
        aqi = ((300 - 201) / (424.0 - 355.0)) * (median_value - 355.0) + 201
    elif 425.0 <= median_value <= 504.0:
        aqi = ((400 - 301) / (504.0 - 425.0)) * (median_value - 425.0) + 301
    elif 505.0 <= median_value <= 604.0:
        aqi = ((500 - 401) / (604.0 - 505.0)) * (median_value - 505.0) + 401
    else:
        aqi = 500  # Beyond the AQI scale

    return round(aqi)

def category_level(aqi):
    category = ""
    if aqi <= 50:
        category = "Good"
    elif aqi <= 100:
        category = "Moderate"
    elif aqi <= 150:
        category = "Unhealthy for Sensitive Group"
    elif aqi <= 200:
        category = "Unhealthy"
    elif aqi <= 300:
        category = "Very Unhealthy"
    else:
        category = "Hazardous"
    return category


# temp = []
temp = {}
temp_tgl = {}
temp_final = {}
for sht in sheet_list:
    low = True
    temp_tgl[sht] = {'tgl':[]}
    temp[sht] = {'25': {'nilai_aqi':[],'level':[]},'10':{'nilai_aqi':[],'level':[]}}
    temp_final[sht] = []
    # print(sht)
    df = pd.read_excel('template_upload.xlsx',sht)
    # print(df)

# df = pd.read_excel('template_upload.xlsx','jakarta')
# print("-----------------------")
    # status_max = False
    for record in df['date']: ## ini dianggap nilai 2.5 dan 10 recordnya sama, jika berdeba perlu penyesuaian filter mana record lebih banyak sebagai acuan looping
        tmp = str(record)
        explode = tmp.split('T')
        temp_tgl[sht]['tgl'].append(explode[0])
    for idx, record in enumerate(df['median_2.5']): ## ini dianggap nilai 2.5 dan 10 recordnya sama, jika berdeba perlu penyesuaian filter mana record lebih banyak sebagai acuan looping
        val = calculate_aqi_2_5(record)
        val_category = category_level(record)
        temp[sht]['25']['nilai_aqi'].append(val)
        temp[sht]['25']['level'].append(val_category)

        # final 
        temp_final[sht].append({'tgl':temp_tgl[sht]['tgl'][idx],'val':{'25':{'nilai_aqi':val,'level':val_category},'10':{},'max':{}}})

    for idx, record in enumerate(df['median_10']): ## ini dianggap nilai 2.5 dan 10 recordnya sama, jika berdeba perlu penyesuaian filter mana record lebih banyak sebagai acuan looping
        val = calculate_aqi_10(record)
        val_category = category_level(record)
        temp[sht]['10']['nilai_aqi'].append(val)
        temp[sht]['10']['level'].append(val_category)

        # final 
        temp_final[sht][idx]['val']['10'] = {'nilai_aqi':val,'level':val_category}
        if(low):
            temp_final[sht][idx]['val']['max'] = 200



# looping data temp untuk membuat report sample pengambilan data
# print(temp)
print(temp_final)
# print(temp['jakarta']['25']['nilai_aqi'])
# print(temp['jakarta']['10']['nilai_aqi'])
# print(temp['jakarta']['10']['level'])



